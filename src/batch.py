import argparse
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from researcher import research
from config import DB_PATH, INPUT_XLSX

# Fresh-pass mode: research every populated data row.
TRUSTED_FIRST_DATA_ROWS = 0

RESEARCH_STATUSES = {'VERIFIED', 'NOT_SERIES', 'LIKELY_NOT_SERIES', 'LIKELY_SERIES', 'UNFOUND', 'CONFLICT', 'TIMED_OUT', 'ERROR'}


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def connect():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA synchronous=NORMAL")
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS jobs (
            row_number INTEGER PRIMARY KEY,
            raw_title TEXT,
            raw_author TEXT,
            publication_year TEXT,
            status TEXT NOT NULL,
            attempts INTEGER NOT NULL DEFAULT 0,
            result_json TEXT,
            last_error TEXT,
            updated_at TEXT NOT NULL
        )
        """
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_jobs_status_row "
        "ON jobs(status, row_number)"
    )
    db.commit()
    return db


def workbook_rows():
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))

    required = [
        "Title",
        "Author",
        "Publication Year",
        "Series Y/N",
        "Number in Series",
        "Name of Series",
        "Source",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(
            "Missing required workbook columns: " + ", ".join(missing)
        )

    idx = {name: headers.index(name) for name in required}

    for row_number, row in enumerate(rows, start=2):
        if not any(v is not None for v in row):
            continue

        def value(name):
            i = idx[name]
            return row[i] if i < len(row) else None

        yield {
            "row_number": row_number,
            "title": value("Title"),
            "author": value("Author"),
            "publication_year": value("Publication Year"),
            "series_yn": value("Series Y/N"),
            "series_number": value("Number in Series"),
            "series_name": value("Name of Series"),
            "source": value("Source"),
        }


def initialize():
    db = connect()
    inserted = 0

    for item in workbook_rows():
        # Excel row 2 is data row 1.
        data_index = item["row_number"] - 1
        is_trusted = data_index <= TRUSTED_FIRST_DATA_ROWS
        status = "PREEXISTING" if is_trusted else "PENDING"

        result_json = None
        if is_trusted:
            existing_source = (
                str(item["source"]).strip()
                if item["source"] is not None
                else ""
            )
            result_json = json.dumps(
                {
                    "title": item["title"],
                    "author": item["author"],
                    "status": "PREEXISTING",
                    "is_series": str(item["series_yn"] or "")
                    .strip()
                    .lower()
                    in {"yes", "y", "true", "1"},
                    "series_name": item["series_name"],
                    "series_number": item["series_number"],
                    "confidence": None,
                    "searches_used": 0,
                    "pages_fetched": 0,
                    "sources": (
                        [{"url": existing_source}] if existing_source else []
                    ),
                    "conflicts": [],
                    "notes": (
                        "Trusted existing workbook research; "
                        "not re-researched."
                    ),
                },
                ensure_ascii=False,
            )

        cur = db.execute(
            """
            INSERT OR IGNORE INTO jobs
            (
                row_number,
                raw_title,
                raw_author,
                publication_year,
                status,
                attempts,
                result_json,
                last_error,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, 0, ?, NULL, ?)
            """,
            (
                item["row_number"],
                None if item["title"] is None else str(item["title"]),
                None if item["author"] is None else str(item["author"]),
                (
                    None
                    if item["publication_year"] is None
                    else str(item["publication_year"])
                ),
                status,
                result_json,
                now_iso(),
            ),
        )
        if cur.rowcount:
            inserted += 1

    # If a previous run died mid-book, make that row resumable.
    db.execute(
        """
        UPDATE jobs
        SET status='PENDING', updated_at=?
        WHERE status='IN_PROGRESS'
        """,
        (now_iso(),),
    )
    db.commit()

    total = db.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
    trusted = db.execute(
        "SELECT COUNT(*) FROM jobs WHERE status='PREEXISTING'"
    ).fetchone()[0]
    pending = db.execute(
        "SELECT COUNT(*) FROM jobs WHERE status='PENDING'"
    ).fetchone()[0]

    print(
        json.dumps(
            {
                "database": str(DB_PATH),
                "total_jobs": total,
                "new_jobs_inserted": inserted,
                "trusted_preexisting": trusted,
                "pending": pending,
            },
            indent=2,
        )
    )
    db.close()


def status():
    db = connect()
    rows = db.execute(
        """
        SELECT status, COUNT(*)
        FROM jobs
        GROUP BY status
        ORDER BY status
        """
    ).fetchall()
    print(json.dumps({key: count for key, count in rows}, indent=2))
    db.close()


def run_batch(limit=None):
    db = connect()

    # Recover any row left in-progress by an interrupted run.
    db.execute(
        """
        UPDATE jobs
        SET status='PENDING', updated_at=?
        WHERE status='IN_PROGRESS'
        """,
        (now_iso(),),
    )
    db.commit()

    processed = 0

    while limit is None or processed < limit:
        row = db.execute(
            """
            SELECT
                row_number,
                raw_title,
                raw_author,
                publication_year
            FROM jobs
            WHERE status='PENDING'
            ORDER BY row_number
            LIMIT 1
            """
        ).fetchone()

        if row is None:
            print("No PENDING rows remain.")
            break

        row_number, title, author, year = row

        db.execute(
            """
            UPDATE jobs
            SET status='IN_PROGRESS',
                attempts=attempts+1,
                updated_at=?
            WHERE row_number=?
            """,
            (now_iso(), row_number),
        )
        db.commit()

        if not title or not author:
            result = {
                "title": title,
                "author": author,
                "status": "UNFOUND",
                "is_series": None,
                "series_name": None,
                "series_number": None,
                "confidence": 0.0,
                "searches_used": 0,
                "pages_fetched": 0,
                "sources": [],
                "conflicts": [],
                "notes": "Missing title or author in source workbook.",
            }
        else:
            try:
                result = research(title, author, year)
            except Exception as exc:
                result = {
                    "title": title,
                    "author": author,
                    "status": "ERROR",
                    "is_series": None,
                    "series_name": None,
                    "series_number": None,
                    "confidence": 0.0,
                    "searches_used": 0,
                    "pages_fetched": 0,
                    "sources": [],
                    "conflicts": [],
                    "notes": "Batch runner exception: " + repr(exc),
                }

        original_status = result.get("status")
        result_status = original_status

        if result_status not in RESEARCH_STATUSES:
            result_status = "ERROR"
            result["status"] = "ERROR"
            result["notes"] = (
                "Invalid result status returned by researcher: "
                + repr(original_status)
            )

        last_error = (
            result.get("notes") if result_status == "ERROR" else None
        )

        db.execute(
            """
            UPDATE jobs
            SET status=?,
                result_json=?,
                last_error=?,
                updated_at=?
            WHERE row_number=?
            """,
            (
                result_status,
                json.dumps(result, ensure_ascii=False),
                last_error,
                now_iso(),
                row_number,
            ),
        )
        db.commit()

        processed += 1

        confidence = result.get("confidence")
        if (
            isinstance(confidence, (int, float))
            and not isinstance(confidence, bool)
        ):
            confidence_text = f"{confidence * 100:.0f}%"
        else:
            confidence_text = "n/a"

        series_name = result.get("series_name")
        series_text = str(series_name).strip() if series_name else "-"

        print(
            f"row={row_number} status={result_status} "
            f"confidence={confidence_text} "
            f"series={series_text!r} "
            f"title={title!r}",
            flush=True,
        )

    db.close()



RETRYABLE_STATUSES = (
    "CONFLICT",
    "UNFOUND",
    "TIMED_OUT",
    "ERROR",
)


def retry_unresolved():
    """Requeue unresolved/failed jobs without touching successful results."""
    db = connect()

    active = db.execute(
        "SELECT row_number, raw_title FROM jobs WHERE status='IN_PROGRESS'"
    ).fetchall()

    if active:
        db.close()
        raise RuntimeError(
            "Cannot requeue unresolved jobs while research is IN_PROGRESS: "
            + repr(active)
        )

    before = {
        status: db.execute(
            "SELECT COUNT(*) FROM jobs WHERE status=?",
            (status,),
        ).fetchone()[0]
        for status in RETRYABLE_STATUSES
    }

    placeholders = ",".join("?" for _ in RETRYABLE_STATUSES)

    cur = db.execute(
        f"""
        UPDATE jobs
        SET status='PENDING',
            updated_at=?
        WHERE status IN ({placeholders})
        """,
        (now_iso(), *RETRYABLE_STATUSES),
    )

    db.commit()

    print(
        json.dumps(
            {
                "requeued": cur.rowcount,
                "previous_status_counts": before,
                "retryable_statuses": list(RETRYABLE_STATUSES),
            },
            indent=2,
        )
    )

    db.close()

def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)

    sub.add_parser("init")
    sub.add_parser("status")
    sub.add_parser(
        "retry",
        help="Requeue CONFLICT, UNFOUND, TIMED_OUT, and ERROR jobs",
    )

    p_run = sub.add_parser("run")
    p_run.add_argument("--limit", type=int)

    args = parser.parse_args()

    if args.command == "init":
        initialize()
    elif args.command == "status":
        status()
    elif args.command == "retry":
        retry_unresolved()
    elif args.command == "run":
        run_batch(args.limit)


if __name__ == "__main__":
    main()
