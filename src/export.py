import argparse
import json
import sqlite3
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from normalize import normalize
from config import DB_PATH, DEFAULT_OUTPUT, INPUT_XLSX

KNOWN_STATUSES = {
    'PENDING',
    'IN_PROGRESS',
    'PREEXISTING',
    'VERIFIED',
    'NOT_SERIES',
    'LIKELY_NOT_SERIES',
    'LIKELY_SERIES',
    'UNFOUND',
    'CONFLICT',
    'TIMED_OUT',
    'ERROR',
}

REVIEW_COLUMNS_BEFORE_SOURCE = [
    'Series Status',
    'Series Total As Of',
    'Series Position Text',
    'Alternate Series Names',
    'ISBN-13',
    'ISBN-10',
    'Publisher',
    'Original Publication Year',
    'Research Status',
    'Confidence',
    'Searches Used',
    'Pages Fetched',
    'Search Seconds',
    'Fetch Seconds',
    'Model Seconds',
    'Total Research Seconds',
    'Search Cache Hit',
    'Researched At',
]

REVIEW_COLUMNS_AFTER_SOURCES = [
    'Notes',
    'Conflicts',
    'Evidence Archive',
]


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def clean_url(value):
    if value is None:
        return ''
    return str(value).strip()


def result_sources(result):
    urls = []
    for source in result.get('sources') or []:
        if isinstance(source, dict):
            url = clean_url(source.get('url'))
        else:
            url = clean_url(source)
        if url and url not in urls:
            urls.append(url)
    return urls


def load_jobs():
    db = sqlite3.connect(DB_PATH)
    rows = db.execute(
        '''
        SELECT
            row_number,
            raw_title,
            raw_author,
            publication_year,
            status,
            result_json,
            last_error
        FROM jobs
        ORDER BY row_number
        '''
    ).fetchall()
    db.close()

    jobs = {}
    max_sources = 1

    for (
        row_number,
        raw_title,
        raw_author,
        publication_year,
        status,
        result_json,
        last_error,
    ) in rows:
        result = {}
        if result_json:
            try:
                result = json.loads(result_json)
            except json.JSONDecodeError:
                result = {
                    'notes': 'Exporter could not parse result_json.',
                }

        sources = result_sources(result)
        max_sources = max(max_sources, len(sources))

        jobs[row_number] = {
            'status': status,
            'result': result,
            'sources': sources,
            'raw_title': raw_title,
            'raw_author': raw_author,
            'publication_year': publication_year,
            'last_error': last_error,
        }

    return jobs, max_sources


def copy_cell_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def header_map(ws):
    return {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def text_or_none(value):
    if value is None:
        return None
    if isinstance(value, (list, tuple, set)):
        items = [str(x).strip() for x in value if str(x).strip()]
        return ' | '.join(items) if items else None
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = str(value).strip()
    return text if text else None


def conflicts_text(value):
    if not value:
        return None
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    except TypeError:
        return str(value)


def set_url_cell(cell, url):
    if not url:
        cell.value = None
        cell.hyperlink = None
        return
    cell.value = url
    cell.hyperlink = url
    cell.style = 'Hyperlink'


def add_summary_sheet(wb, counts, total_jobs, max_sources, output_path):
    if 'Research Summary' in wb.sheetnames:
        del wb['Research Summary']

    ws = wb.create_sheet('Research Summary', 0)
    ws['A1'] = 'Iris Library Research Export'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A3'] = 'Generated At (UTC)'
    ws['B3'] = now_iso()
    ws['A4'] = 'Rows in Research DB'
    ws['B4'] = total_jobs
    ws['A5'] = 'Source Columns'
    ws['B5'] = max_sources
    ws['A6'] = 'Output'
    ws['B6'] = str(output_path)

    ws['A8'] = 'Research Status'
    ws['B8'] = 'Count'
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)

    row = 9
    for status in [
        'VERIFIED',
        'NOT_SERIES',
        'LIKELY_SERIES',
        'LIKELY_NOT_SERIES',
        'CONFLICT',
        'UNFOUND',
        'TIMED_OUT',
        'ERROR',
        'IN_PROGRESS',
        'PENDING',
        'PREEXISTING',
    ]:
        if status in counts:
            ws.cell(row, 1).value = status
            ws.cell(row, 2).value = counts[status]
            row += 1

    terminal = {
        'VERIFIED', 'NOT_SERIES', 'LIKELY_SERIES', 'LIKELY_NOT_SERIES',
        'CONFLICT', 'UNFOUND', 'TIMED_OUT', 'ERROR', 'PREEXISTING'
    }
    completed = sum(count for status, count in counts.items() if status in terminal)
    pending = counts.get('PENDING', 0) + counts.get('IN_PROGRESS', 0)

    ws.cell(row + 1, 1).value = 'Completed / terminal'
    ws.cell(row + 1, 2).value = completed
    ws.cell(row + 2, 1).value = 'Pending / in progress'
    ws.cell(row + 2, 2).value = pending

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 70
    ws.freeze_panes = 'A8'


def export(output_path):
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f'Input workbook not found: {INPUT_XLSX}')
    if not DB_PATH.exists():
        raise FileNotFoundError(f'SQLite database not found: {DB_PATH}')

    jobs, max_sources = load_jobs()

    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required = [
        'Title',
        'Author',
        'Publication Year',
        'Series Y/N',
        'Number in Series',
        'Total Volumes in Series',
        'Name of Series',
        'Source',
    ]
    missing = [h for h in required if h not in headers]
    if missing:
        raise RuntimeError('Missing required workbook columns: ' + ', '.join(missing))

    # Add normalized title immediately after Title.
    title_col = headers.index('Title') + 1
    ws.insert_cols(title_col + 1, 1)
    ws.cell(1, title_col + 1).value = 'Normalized Title'
    copy_cell_style(ws.cell(1, title_col), ws.cell(1, title_col + 1))

    # Add normalized author immediately after Author.
    cols = header_map(ws)
    author_col = cols['Author']
    ws.insert_cols(author_col + 1, 1)
    ws.cell(1, author_col + 1).value = 'Normalized Author'
    copy_cell_style(ws.cell(1, author_col), ws.cell(1, author_col + 1))

    # Insert current research/bonus metadata immediately before the original Source.
    cols = header_map(ws)
    source_col = cols['Source']
    ws.insert_cols(source_col, len(REVIEW_COLUMNS_BEFORE_SOURCE))
    style_header = ws.cell(1, max(1, source_col - 1))
    for offset, name in enumerate(REVIEW_COLUMNS_BEFORE_SOURCE):
        c = source_col + offset
        ws.cell(1, c).value = name
        copy_cell_style(style_header, ws.cell(1, c))
        ws.cell(1, c).alignment = copy(style_header.alignment)
        ws.cell(1, c).alignment = Alignment(
            horizontal=ws.cell(1, c).alignment.horizontal,
            vertical=ws.cell(1, c).alignment.vertical,
            wrap_text=True,
        )

    # Rename original Source to Source 1 and add additional dynamic source columns.
    cols = header_map(ws)
    original_source_col = cols['Source']
    ws.cell(1, original_source_col).value = 'Source 1'
    source1_col = original_source_col

    if max_sources > 1:
        ws.insert_cols(source1_col + 1, max_sources - 1)
        for offset in range(1, max_sources):
            c = source1_col + offset
            ws.cell(1, c).value = f'Source {offset + 1}'
            copy_cell_style(ws.cell(1, source1_col), ws.cell(1, c))

    # Append Notes/Conflicts/Evidence Archive after the source columns.
    after_source_col = source1_col + max_sources
    ws.insert_cols(after_source_col, len(REVIEW_COLUMNS_AFTER_SOURCES))
    for offset, name in enumerate(REVIEW_COLUMNS_AFTER_SOURCES):
        c = after_source_col + offset
        ws.cell(1, c).value = name
        copy_cell_style(ws.cell(1, source1_col), ws.cell(1, c))
        ws.cell(1, c).alignment = Alignment(wrap_text=True)

    cols = header_map(ws)
    source_cols = [cols[f'Source {i}'] for i in range(1, max_sources + 1)]

    # Normalize every populated source row, including PENDING rows.
    for row_number in range(2, ws.max_row + 1):
        raw_title = ws.cell(row_number, cols['Title']).value
        raw_author = ws.cell(row_number, cols['Author']).value
        if raw_title is None and raw_author is None:
            continue

        publication_year = ws.cell(row_number, cols['Publication Year']).value
        normalized = normalize(raw_title, raw_author, publication_year)
        ws.cell(row_number, cols['Normalized Title']).value = normalized.get('search_title')
        ws.cell(row_number, cols['Normalized Author']).value = normalized.get('search_author')

    counts = {}

    # Research-derived fields that must never display stale workbook values
    # during a fresh-pass run.
    clear_headers = [
        'Series Y/N',
        'Number in Series',
        'Total Volumes in Series',
        'Name of Series',
        *REVIEW_COLUMNS_BEFORE_SOURCE,
        *REVIEW_COLUMNS_AFTER_SOURCES,
    ]

    for row_number, job in jobs.items():
        if row_number > ws.max_row:
            continue

        status = job['status']
        result = job['result']
        sources = job['sources']
        bonus = result.get('bonus_metadata')
        if not isinstance(bonus, dict):
            bonus = {}
        timings = result.get('timings')
        if not isinstance(timings, dict):
            timings = {}
        live = result.get('live_research')
        if not isinstance(live, dict):
            live = {}

        counts[status] = counts.get(status, 0) + 1

        if status == 'PREEXISTING':
            ws.cell(row_number, cols['Research Status']).value = status
            continue

        # Clear stale research fields from the source workbook for every
        # fresh-pass row, including PENDING/IN_PROGRESS/failed rows.
        for name in clear_headers:
            if name in cols:
                ws.cell(row_number, cols[name]).value = None
        for c in source_cols:
            set_url_cell(ws.cell(row_number, c), None)

        ws.cell(row_number, cols['Research Status']).value = status

        # Strict final answers only populate Series Y/N.
        if status == 'VERIFIED':
            ws.cell(row_number, cols['Series Y/N']).value = 'Yes'
            ws.cell(row_number, cols['Number in Series']).value = result.get('series_number')
            ws.cell(row_number, cols['Name of Series']).value = result.get('series_name')
        elif status == 'NOT_SERIES':
            ws.cell(row_number, cols['Series Y/N']).value = 'No'
        elif status == 'LIKELY_SERIES':
            # Preserve useful candidate data, but deliberately leave Series Y/N blank
            # so a provisional conclusion cannot masquerade as VERIFIED.
            ws.cell(row_number, cols['Number in Series']).value = result.get('series_number')
            ws.cell(row_number, cols['Name of Series']).value = result.get('series_name')

        # Bonus metadata is reviewable for any completed result that contains it.
        ws.cell(row_number, cols['Total Volumes in Series']).value = bonus.get('total_volumes_in_series')
        ws.cell(row_number, cols['Series Status']).value = bonus.get('series_status')
        ws.cell(row_number, cols['Series Total As Of']).value = bonus.get('series_total_as_of')
        ws.cell(row_number, cols['Series Position Text']).value = text_or_none(bonus.get('series_position_text'))
        ws.cell(row_number, cols['Alternate Series Names']).value = text_or_none(bonus.get('alternate_series_names'))
        ws.cell(row_number, cols['ISBN-13']).value = text_or_none(bonus.get('isbn_13'))
        ws.cell(row_number, cols['ISBN-10']).value = text_or_none(bonus.get('isbn_10'))
        ws.cell(row_number, cols['Publisher']).value = text_or_none(bonus.get('publisher'))
        ws.cell(row_number, cols['Original Publication Year']).value = bonus.get('original_publication_year')

        confidence = result.get('confidence')
        if isinstance(confidence, (int, float)) and not isinstance(confidence, bool):
            ws.cell(row_number, cols['Confidence']).value = float(confidence)
            ws.cell(row_number, cols['Confidence']).number_format = '0%'

        ws.cell(row_number, cols['Searches Used']).value = result.get('searches_used')
        ws.cell(row_number, cols['Pages Fetched']).value = result.get('pages_fetched')
        ws.cell(row_number, cols['Search Seconds']).value = timings.get('search_seconds')
        ws.cell(row_number, cols['Fetch Seconds']).value = timings.get('fetch_seconds')
        ws.cell(row_number, cols['Model Seconds']).value = timings.get('model_seconds')
        ws.cell(row_number, cols['Total Research Seconds']).value = timings.get('total_seconds')
        ws.cell(row_number, cols['Search Cache Hit']).value = live.get('search_cache_hit')
        ws.cell(row_number, cols['Researched At']).value = result.get('researched_at')

        for i, url in enumerate(sources):
            if i >= len(source_cols):
                break
            set_url_cell(ws.cell(row_number, source_cols[i]), url)

        notes = result.get('notes')
        if notes is None and status == 'ERROR':
            notes = job.get('last_error')
        ws.cell(row_number, cols['Notes']).value = text_or_none(notes)
        ws.cell(row_number, cols['Conflicts']).value = conflicts_text(result.get('conflicts'))
        ws.cell(row_number, cols['Evidence Archive']).value = text_or_none(result.get('evidence_archive'))

    # Jobs table should normally cover every workbook row. If not, mark the
    # workbook-only rows explicitly instead of silently retaining stale research.
    for row_number in range(2, ws.max_row + 1):
        if row_number in jobs:
            continue
        title = ws.cell(row_number, cols['Title']).value
        author = ws.cell(row_number, cols['Author']).value
        if title is None and author is None:
            continue
        for name in clear_headers:
            if name in cols:
                ws.cell(row_number, cols[name]).value = None
        for c in source_cols:
            set_url_cell(ws.cell(row_number, c), None)
        ws.cell(row_number, cols['Research Status']).value = 'NOT_IN_DB'
        counts['NOT_IN_DB'] = counts.get('NOT_IN_DB', 0) + 1

    # Review-friendly widths.
    widths = {
        'Normalized Title': 38,
        'Normalized Author': 28,
        'Research Status': 20,
        'Series Status': 15,
        'Series Total As Of': 16,
        'Series Position Text': 28,
        'Alternate Series Names': 36,
        'ISBN-13': 18,
        'ISBN-10': 16,
        'Publisher': 30,
        'Original Publication Year': 20,
        'Confidence': 12,
        'Searches Used': 14,
        'Pages Fetched': 14,
        'Search Seconds': 14,
        'Fetch Seconds': 14,
        'Model Seconds': 14,
        'Total Research Seconds': 18,
        'Search Cache Hit': 16,
        'Researched At': 26,
        'Notes': 60,
        'Conflicts': 45,
        'Evidence Archive': 45,
    }
    for name, width in widths.items():
        if name in cols:
            ws.column_dimensions[ws.cell(1, cols[name]).column_letter].width = width

    for c in source_cols:
        ws.column_dimensions[ws.cell(1, c).column_letter].width = 45

    # Wrap long review fields.
    wrap_headers = {
        'Series Position Text', 'Alternate Series Names', 'Notes', 'Conflicts',
        'Evidence Archive', *[f'Source {i}' for i in range(1, max_sources + 1)]
    }
    for name in wrap_headers:
        if name not in cols:
            continue
        c = cols[name]
        for row_number in range(1, ws.max_row + 1):
            cell = ws.cell(row_number, c)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical='top',
                wrap_text=True,
            )

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    add_summary_sheet(wb, counts, len(jobs), max_sources, output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(
        json.dumps(
            {
                'exporter_version': 'final-review-v1',
                'output': str(output_path),
                'rows_in_database': len(jobs),
                'status_counts': counts,
                'source_columns': max_sources,
                'bonus_metadata_columns': [
                    'Total Volumes in Series',
                    'Series Status',
                    'Series Total As Of',
                    'Series Position Text',
                    'Alternate Series Names',
                    'ISBN-13',
                    'ISBN-10',
                    'Publisher',
                    'Original Publication Year',
                ],
                'review_columns': [
                    'Research Status',
                    'Confidence',
                    'Searches Used',
                    'Pages Fetched',
                    'Search Seconds',
                    'Fetch Seconds',
                    'Model Seconds',
                    'Total Research Seconds',
                    'Search Cache Hit',
                    'Researched At',
                    'Notes',
                    'Conflicts',
                    'Evidence Archive',
                ],
                'input_untouched': str(INPUT_XLSX),
            },
            indent=2,
        )
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--output',
        default=str(DEFAULT_OUTPUT),
        help='Output .xlsx path',
    )
    args = parser.parse_args()
    export(Path(args.output))


if __name__ == '__main__':
    main()
